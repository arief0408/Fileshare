import sys
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

def write_text_to_excel(excel_file, sheet_name, row, col, text):
    # Check if the Excel file exists
    if os.path.exists(excel_file):
        # Load the existing workbook
        wb = load_workbook(excel_file)
        # If the sheet doesn't exist, create it
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        sheet = wb[sheet_name]
    else:
        # Create a new workbook and add the specified sheet
        wb = Workbook()
        sheet = wb.active
        sheet.title = sheet_name

    # Write the text into the specified cell
    cell = sheet[f"{col}{row}"]
    cell.value = text

    # Set the column width to 80
    sheet.column_dimensions[col].width = 100

    # Set the font to Courier New
    cell.font = Font(name='Courier New', size=12)

    # Set text to wrap automatically
    cell.alignment = Alignment(wrap_text=True)

    # Save the updated or new Excel file
    try:
        wb.save(excel_file)
        print(f"Successfully saved the file: {excel_file}")
    except PermissionError:
        print(f"Error: Permission denied. Please ensure that '{excel_file}' is closed.")
        sys.exit(1)

if __name__ == "__main__":
    # Check the number of arguments
    if len(sys.argv) < 6:
        print("Usage: python write_excel.py <excel_file> <sheet_name> <row> <col> <text...>")
        sys.exit(1)
    excel_file = sys.argv[1]   # Excel File
    sheet_name = sys.argv[2]  # Sheet name
    row = sys.argv[3]          # Row number
    col = sys.argv[4]          # Column letter
    
    
    # Capture the entire text as a single string, preserving spaces
    text = " ".join(sys.argv[5:])  # Join arguments without altering spaces

    # Optionally replace single spaces with double spaces to ensure original spacing is visible
    text = text.replace(" ", "  ")

    write_text_to_excel(excel_file, sheet_name, row, col, text)
    print(f"Text written into {sheet_name} at {col}{row}, with Courier New font, column width set to 80, and text wrapping enabled.")
