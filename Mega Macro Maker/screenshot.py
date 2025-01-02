import sys
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image
import pyautogui  # For taking the screenshot
import openpyxl.utils  # For handling column index to letter conversion

def take_screenshot():
    # Take a screenshot
    screenshot = pyautogui.screenshot()
    return screenshot

def resize_image(image, scale=0.5):
    # Resize the image to a specified scale
    width, height = image.size
    new_size = (int(width * scale), int(height * scale))
    resized_image = image.resize(new_size)
    return resized_image

def save_image_to_excel(excel_file, sheet_name, row, image):
    # Check if the Excel file exists
    if os.path.exists(excel_file):
        # Load the existing workbook
        wb = load_workbook(excel_file)
        # If the sheet doesn't exist, create it
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        sheet = wb[sheet_name]
    else:
        # Create a new workbook and add the specified sheet
        wb = Workbook()
        sheet = wb.active
        sheet.title = sheet_name

    # Find the first empty column in the specified row
    col_index = 1
    while sheet.cell(row=row, column=col_index).value is not None:
        col_index += 1
    col = openpyxl.utils.get_column_letter(col_index)  # Convert column index to letter

    # Save the image to a temporary file
    temp_image_path = "screenshot.png"
    image.save(temp_image_path)

    # Insert the image into the specified cell
    img = Image(temp_image_path)
    sheet.add_image(img, f"{col}{row}")

    # Set the column width and row height to match the image size
    img_width, img_height = image.size
    sheet.column_dimensions[col].width = img_width / 7  # Adjust width factor as needed
    sheet.row_dimensions[int(row)].height = 295  # Set the row height to 295 for image

    # Save the updated or new Excel file
    try:
        wb.save(excel_file)
        #print(f"Successfully saved the file: {excel_file}")
    except PermissionError:
        #print(f"Error: Permission denied. Please ensure that '{excel_file}' is closed.")
        sys.exit(1)

def write_text_to_excel(excel_file, sheet_name, row, text):
    # Check if the Excel file exists
    if os.path.exists(excel_file):
        # Load the existing workbook
        wb = load_workbook(excel_file)
        # If the sheet doesn't exist, create it
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        sheet = wb[sheet_name]
    else:
        # Create a new workbook and add the specified sheet
        wb = Workbook()
        sheet = wb.active
        sheet.title = sheet_name

    # Find the first empty column in the specified row
    col_index = 1
    while sheet.cell(row=row, column=col_index).value is not None:
        col_index += 1
    col = openpyxl.utils.get_column_letter(col_index)  # Convert column index to letter

    # Write the text into the specified cell
    sheet[f"{col}{row}"].value = text

    # Set the column width to 80 for text content
    sheet.column_dimensions[col].width = 10

    # Set the font to Courier New
    sheet[f"{col}{row}"].font = Font(name='Courier New', size=12)

    # Set text to wrap automatically
    sheet[f"{col}{row}"].alignment = Alignment(wrap_text=True)

    # Set the row height to a smaller value for text content
    sheet.row_dimensions[int(row)].height = 15

    # Save the updated or new Excel file
    try:
        wb.save(excel_file)
        #print(f"Successfully saved the file: {excel_file}")
    except PermissionError:
        #print(f"Error: Permission denied. Please ensure that '{excel_file}' is closed.")
        sys.exit(1)

if __name__ == "__main__":
    # Check the number of arguments
    if len(sys.argv) < 5:
        #print("Usage: python combined_script.py <excel_file> <sheet_name> <row> <text...>")
        sys.exit(1)

    excel_file = sys.argv[1]   # Excel File
    sheet_name = sys.argv[2]    # Sheet name
    row = sys.argv[3]           # Row number

    # Capture the entire text as a single string, preserving spaces
    text = " ".join(sys.argv[4:])  # Join arguments without altering spaces

    # Check if the text is "SCREENSHOT"
    if text.strip().upper() == "SCREENSHOT":
        # Take a screenshot
        screenshot = take_screenshot()
        
        # Resize the screenshot to half its size
        resized_screenshot = resize_image(screenshot)
        
        # Save the resized image to Excel
        save_image_to_excel(excel_file, sheet_name, row, resized_screenshot)
        #print(f"Screenshot inserted into {sheet_name} at row {row} with size adjusted.")
    else:
        # Optionally replace single spaces with double spaces to ensure original spacing is visible
        text = text.replace(" ", "  ")
        write_text_to_excel(excel_file, sheet_name, row, text)
        #print(f"Text written into {sheet_name} at row {row}, with Courier New font, column width set to 80, and text wrapping enabled.")