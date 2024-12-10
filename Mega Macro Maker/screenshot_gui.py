import tkinter as tk
import pyautogui
import openpyxl
from PIL import Image
from io import BytesIO
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
import time

# Track the current row and column for image placement
current_row = 1
current_col = 1

def find_next_available_cell(ws):
    """Find the next empty cell in the worksheet."""
    global current_row, current_col

    while True:
        col_letter = get_column_letter(current_col)
        cell = ws[f"{col_letter}{current_row}"]
        if cell.value is None:  # Check if the cell is empty
            return col_letter, current_row
        current_col += 1
        if current_col > 26:  # If beyond column Z, move to the next row
            current_col = 1
            current_row += 1

def take_screenshot_and_insert():
    global current_row, current_col

    # Hide the GUI window and update it to ensure it's hidden
    root.withdraw()
    root.update_idletasks()
    time.sleep(0.5)  # Add a slight delay to ensure the GUI is hidden

    # Get the sheet name from the entry box
    sheet_name = sheet_name_entry.get()

    # Take a screenshot
    screenshot = pyautogui.screenshot()

    # Resize the screenshot to 13 cm x 23 cm
    dpi = 96  # Standard screen DPI
    width_pixels = int(23 * dpi / 2.54)  # Convert 23 cm to pixels
    height_pixels = int(13 * dpi / 2.54)  # Convert 13 cm to pixels
    resized_screenshot = screenshot.resize((width_pixels, height_pixels), Image.LANCZOS)

    # Save resized screenshot to memory as a BytesIO object
    image_bytes = BytesIO()
    resized_screenshot.save(image_bytes, format='PNG')
    image_bytes.seek(0)

    # Load or create an Excel workbook
    try:
        wb = openpyxl.load_workbook('screenshot.xlsx')
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    # Check if the sheet exists; if not, create a new one
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    # Determine the next available cell position
    col_letter, row = find_next_available_cell(ws)
    cell = f"{col_letter}{row}"

    # Create an image object for Excel
    img = ExcelImage(image_bytes)

    # Add the image to the worksheet
    ws.add_image(img, cell)

    # Adjust the column width and row height to match the image size
    column_width = width_pixels / 7.5  # Approximation for Excel column width
    row_height = height_pixels * 0.75  # Approximation for Excel row height in points

    ws.column_dimensions[col_letter].width = column_width
    ws.row_dimensions[row].height = row_height

    # Save the Excel file
    wb.save('screenshot.xlsx')

    # Update current_row and current_col for next image placement
    current_col += 1
    if current_col > 26:  # If beyond column Z, move to the next row
        current_col = 1
        current_row += 1

    # Re-show the GUI
    root.deiconify()

# Create the main window (GUI)
root = tk.Tk()
root.title("Screenshot to Excel")

# Set the window size
root.geometry("300x150")

# Label and input for sheet name
sheet_name_label = tk.Label(root, text="Enter Sheet Name:")
sheet_name_label.pack(pady=5)

sheet_name_entry = tk.Entry(root)
sheet_name_entry.pack(pady=5)

# Create a button to trigger the screenshot and insertion
screenshot_button = tk.Button(root, text="Take Screenshot", command=take_screenshot_and_insert)
screenshot_button.pack(pady=20)

# Start the GUI loop
root.mainloop()
