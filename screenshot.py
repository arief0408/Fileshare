import sys
import os
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
import pyautogui  # For taking the screenshot
from PIL import Image as PILImage

def take_screenshot():
    # Take a screenshot
    screenshot = pyautogui.screenshot()
    return screenshot

def resize_image(image, scale=0.5):
    # Resize the image to a specified scale
    width, height = image.size
    new_size = (int(width * scale), int(height * scale))
    resized_image = image.resize(new_size)
    return resized_image

def save_image_to_excel(excel_file, sheet_name, row, col, image):
    # Check if the Excel file exists
    if os.path.exists(excel_file):
        # Load the existing workbook
        wb = load_workbook(excel_file)
        # If the sheet doesn't exist, create it
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        sheet = wb[sheet_name]
    else:
        # Create a new workbook and add the specified sheet
        wb = Workbook()
        sheet = wb.active
        sheet.title = sheet_name

    # Save the image to a temporary file
    temp_image_path = "screenshot.png"
    image.save(temp_image_path)

    # Insert the image into the specified cell
    img = Image(temp_image_path)
    sheet.add_image(img, f"{col}{row}")

    # Set the column width and row height to match the image size
    img_width, img_height = image.size
    sheet.column_dimensions[col].width = img_width / 7  # Adjust width factor as needed
    sheet.row_dimensions[int(row)].height = img_height / 0.75  # Adjust height factor as needed

    # Save the updated or new Excel file
    wb.save(excel_file)

    # Remove the temporary image file
    if os.path.exists(temp_image_path):
        os.remove(temp_image_path)

if __name__ == "__main__":
    # Check the number of arguments
    if len(sys.argv) != 5:
        print("Usage: python screenshot.py <excel_file> <sheet_name> <row> <col>")
        sys.exit(1)

    # Get the command line arguments
    excel_file = sys.argv[1]  # Excel file name
    sheet_name = sys.argv[2]   # Sheet name
    row = sys.argv[3]          # Row number (string)
    col = sys.argv[4]          # Column letter

    # Take a screenshot
    screenshot = take_screenshot()
    
    # Resize the screenshot to half its size
    resized_screenshot = resize_image(screenshot)

    # Save the resized image to Excel
    save_image_to_excel(excel_file, sheet_name, row, col, resized_screenshot)
    print(f"Screenshot inserted into {sheet_name} at {col}{row} with size adjusted.")
